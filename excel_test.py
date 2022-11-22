from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Estoque.xlsx')
ws = wb.active

for row in range(1,11):
    for col in range(1,6):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

ws.insert_rows(4)

ws.delete_rows(4)


wb.save('Estoque.xlsx')