from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlwings
import pandas
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

#Abrindo a planilha e iniciando variavel de quantidade de linhas já usadas no Estoque
wb = load_workbook('Estoque.xlsx')
ws = wb["SmartHaus"]

end_row = len(ws["A"])

#Iniciando classe de objeto para organizar as informações de cada produto
class Produtos:
    def __init__(self,nome,qnt,valor,descricao,marca):
        self.nome = nome
        self.qnt = qnt
        self.valor = valor
        self.descricao = descricao
        self.marca = marca

#Adicionando novos produtos ao banco de dados
def adicionar_estoque():
    wb = load_workbook('Estoque.xlsx')
    ws = wb["SmartHaus"]
    while True:
        nome = input("Nome do produto:")
        if len(nome) != 0:
            break
        else:
            print("Campo obrigatório")
    while True:
        qnt = input("Quantos estão entrando no estoque?")
        if qnt.isdigit():
            break
        elif len(qnt) == 0:
            qnt = 0
            break
        else:
            print("Quantidade deve ser números")
    while True:
        valor = input("Valor por unidade:")
        if valor.isdigit():
            valor = int(valor)
            if valor > 0:
                break
            else:
                print("O valor não pode ser 0")
        else:
            print("O valor deve ser número")
    while True:
        descricao = input("Descrição do produto:")
        if len(descricao) == 0:
            descricao = "Produto sem descrição."
            break
        else:
            break
    while True:
        marca = input("Marca do produto:")
        if len(marca) != 0:
            break
        else:
            print("Campo obrigatório")
    end_row = len(ws["A"]) + 1
    ws["A" + str(end_row)] = nome
    ws["B" + str(end_row)] = qnt
    ws["C" + str(end_row)] = valor
    ws["D" + str(end_row)] = descricao
    ws["E" + str(end_row)] = marca
    print("Produto adicionado ao estoque!")
    wb.save('Estoque.xlsx')
    
#Adicionar qnt de um produto já existente
def adicionar_qnt():
    produto = input("Qual produto gostaria de dar entrada?")
    produto = str(produto)
    wb = load_workbook('Estoque.xlsx')
    ws = wb["SmartHaus"]
    for row in range(2,end_row+1):
        if ws["A"+str(row)].value == produto:
            while True:
                quantidade = input("Quantidade:")
                if quantidade.isdigit():
                    quantidade = int(quantidade)
                    if quantidade >= 0:
                        ws["B"+str(row)] = int(ws["B"+str(row)].value) + quantidade
                        wb.save('Estoque.xlsx')
                        return print("Produtos adicionados ao estoque!")
                    else:
                        print("A quantidade não pode ser zero ou valor negativo.")
                    ws["B"+str(row)] = int(ws["B"+str(row)].value) + quantidade
                    wb.save('Estoque.xlsx')
                    return print("Produtos adicionados ao estoque!")
                else:
                    print("Quantidade a ser inserida deve ser um número.")
        
    print("Produto não encontrado no estoque.")

#Subtrai qnt de um produto já existente
def retirar_prod():
    produto = input("Qual produto gostaria de retirar?")
    produto = str(produto)
    wb = load_workbook('Estoque.xlsx')
    ws = wb["SmartHaus"]
    for row in range(2,end_row+1):
        if ws["A"+str(row)].value == produto:
            while True:
                quantidade = input("Quantidade:")
                if quantidade.isdigit():
                    quantidade = int(quantidade)
                    if quantidade >= 0:
                        if quantidade <= ws["B" + str(row)].value:
                            ws["B"+str(row)] = int(ws["B"+str(row)].value) - quantidade
                            wb.save('Estoque.xlsx')
                            return print("Produtos retirados do estoque!")
                        else:
                            return print("Não é possível retirar. Atualmente existem {} unidades no estoque.".format(str(ws["B"+str(row)].value)))
                    else:
                         print("A quantidade não pode ser zero ou negativo.")
                else:
                    print("Quantidade a ser inserida deve ser um número.")
        
    print("Produto não encontrado no estoque.")

#Buscar lista de produtos reservados
def reserved_list(cliente):
    wb = load_workbook('Estoque.xlsx')
    ws = wb["SmartHaus"]
    end_col = len(ws["1"])
    for column in range(5,end_col+1):
        if ws[get_column_letter(column)+"1"].value == cliente:
            for produto in range(2,end_row+1):
                if ws[str(get_column_letter(column)+str(produto))].value is not None:
                    print("Existem {} de {} reservado para o {}".format(ws["B"+str(produto)].value,ws["A"+str(produto)].value,cliente))
            return print("Estes são os itens reservados") 
    print("Falha")

#Iniciando a janela de interface

janela = Tk()
janela.title("Gerenciamento Galpão")
janela.geometry("670x370")
janela.configure(bg="#ececec")

ttk.Separator(janela, orient=HORIZONTAL).grid(row=0,columnspan=1,ipadx=272)

style = ttk.Style(janela)
style.theme_use("clam")

janela.after(1, lambda: janela.focus_force())

#Inserindo os elementos para interação

tree = ttk.Treeview(janela, columns=("0","1","2","3"),show="headings")
tree.heading("0", text="ID")
tree.column("0",anchor="center", stretch="no", width=30)
tree.heading("1", text="Nome do Produto")
tree.column("1",anchor="center", stretch="no", width=200)
tree.heading("2", text="Qnt")
tree.column("2",anchor="center", stretch="no", width=50)
tree.heading("3", text="Descrição")
tree.column("3", stretch="no", width=320)
tree.place(x=30,y=100)

entrada = Entry(janela, width=100)
entrada.focus_set()
entrada.place(x=30, y=70)

b_create = Button(janela, text="Novo Produto", relief="raised", overrelief=RIDGE, anchor=NW, font=("Verdana 12"), bg="#D3D4D3", fg="#080808")
b_create.place(x=30,y=20)

b_alt_desc = Button(janela, text="Alterar Descrição", relief="raised", overrelief=RIDGE, anchor=NW, font=("Verdana 12"), bg="#D3D4D3", fg="#080808")
b_alt_desc.place(x=170,y=20)

b_update = Button(janela, text="Entrada/Saída", relief="raised", overrelief=RIDGE, anchor=NW, font=("Verdana 12"), bg="#D3D4D3", fg="#080808")
b_update.place(x=340,y=20)

b_delete = Button(janela, text="Deletar Produto", relief="raised", overrelief=RIDGE, anchor=NW, font=("Verdana 12"), bg="#D3D4D3", fg="#080808")
b_delete.place(x=490,y=20)

b_importar = Button(janela, text="Importar Dados CSV", relief="raised", overrelief=RIDGE, anchor=NW, font=("Verdana 12"), bg="#D3D4D3", fg="#080808")
b_importar.place(x=475,y=332)

janela.mainloop()
